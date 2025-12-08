from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional

from openpyxl import load_workbook


# RTL 地区列表（中东、阿拉伯语地区、希伯来语地区等）
RTL_REGIONS = {'MECA', 'ARAB', 'ARABIC', 'SA', 'UAE', 'EG', 'IL', 'ISRAEL', 'JO', 'LB', 'IQ', 'SY'}


def is_rtl_region(region_code: str) -> bool:
    """
    根据地区代码判断是否为 RTL（从右到左）语言
    """
    if not region_code:
        return False
    
    upper_code = region_code.upper()
    # 检查是否包含任何 RTL 地区标识
    return any(rtl in upper_code for rtl in RTL_REGIONS)


def clean_text(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    # Strip Excel formula wrappers
    if s.startswith('="') and s.endswith('"'):
        s = s[2:-1]
    if s.startswith("='") and s.endswith("'"):
        s = s[2:-1]
    if s.startswith("="):
        s = s[1:]
    return s.replace("\r\n", "\n").replace("\r", "\n").strip()


def build_merge_index(ws) -> Dict[Tuple[int, int], Tuple[int, int, int, int]]:
    idx = {}
    for mr in ws.merged_cells.ranges:
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                idx[(r, c)] = (mr.min_row, mr.min_col, mr.max_row, mr.max_col)
    return idx


def get_cell_info(ws, r: int, c: int, merge_idx):
    """获取单元格的值和完整格式信息（文本、对齐、加粗、斜体、颜色）"""
    if (r, c) in merge_idx:
        r0, c0, _, _ = merge_idx[(r, c)]
        cell = ws.cell(r0, c0)
    else:
        cell = ws.cell(r, c)
    
    # 读取单元格的值
    value = cell.value
    
    # 检查是否为百分比格式
    if isinstance(value, (int, float)) and cell.number_format:
        if '%' in cell.number_format:
            value = f"{value * 100:.10g}%"
    
    # 读取对齐信息
    alignment = None
    if cell.alignment and cell.alignment.horizontal:
        alignment = cell.alignment.horizontal  # 'center', 'left', 'right'
    
    # 读取字体信息
    bold = False
    italic = False
    color = None
    
    if cell.font:
        bold = bool(cell.font.bold)
        italic = bool(cell.font.italic)
        
        # 读取字体颜色
        if cell.font.color:
            if hasattr(cell.font.color, 'rgb') and cell.font.color.rgb:
                # RGB 格式：可能是字符串 'AARRGGBB'/'RRGGBB' 或 RGB 对象
                rgb = cell.font.color.rgb
                # 将 RGB 对象转换为字符串
                if not isinstance(rgb, str):
                    rgb = str(rgb)
                # 去掉可能的前缀和空格
                rgb = rgb.strip()
                if len(rgb) == 8:  # ARGB 格式
                    color = f"#{rgb[2:]}"
                elif len(rgb) == 6:  # RGB 格式
                    color = f"#{rgb}"
    
    return {
        'value': value,
        'alignment': alignment,
        'bold': bold,
        'italic': italic,
        'color': color,
        # 保留向后兼容的字段
        'center': alignment == 'center' if alignment else False
    }


def get_value(ws, r: int, c: int, merge_idx):
    """获取单元格的值（仅文本，不含格式）"""
    info = get_cell_info(ws, r, c, merge_idx)
    return info['value']


def row_region_values(ws, r: int, c1: int, c2: int, merge_idx):
    return [clean_text(get_value(ws, r, c, merge_idx)) for c in range(c1, c2 + 1)]


def row_region_values_with_format(ws, r: int, c1: int, c2: int, merge_idx):
    """获取一行的值和格式信息"""
    result = []
    for c in range(c1, c2 + 1):
        info = get_cell_info(ws, r, c, merge_idx)
        result.append({
            'value': clean_text(info['value']),
            'bold': info['bold'],
            'center': info['center']
        })
    return result


def is_row_blank(vals: List[str]) -> bool:
    return all(v == "" for v in vals)


def find_regions(ws, merge_idx):
    regions = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = clean_text(get_value(ws, r, c, merge_idx))
            if v.startswith("REGION-"):
                if (r, c) in merge_idx:
                    r0, c0, _, c1 = merge_idx[(r, c)]
                else:
                    r0, c0, c1 = r, c, c
                regions.append({"code": v.split("REGION-")[-1].strip(), "row": r0, "col_start": c0, "col_end": c1})
    # dedup + sort by start col then row
    seen = set()
    uniq = []
    for rg in regions:
        key = (rg["code"], rg["row"], rg["col_start"], rg["col_end"])
        if key not in seen:
            uniq.append(rg)
            seen.add(key)
    uniq.sort(key=lambda x: (x["col_start"], x["row"]))
    return uniq


def scan_titles(ws, region, merge_idx):
    """
    扫描区域内的所有 TITLE 标记，提取标题类型和右侧的实际标题文本。

    返回格式: [(title_type, title_text, row), ...]
    例如: [("规则", "活动规则说明", 10), ("奖励", "真爱树榜单奖励", 25)]
    """
    c1, c2 = region["col_start"], region["col_end"]
    start_row = region["row"] + 1
    titles = []
    for r in range(start_row, ws.max_row + 1):
        row_vals = row_region_values(ws, r, c1, c2, merge_idx)
        for i, v in enumerate(row_vals):
            if v.startswith("TITLE-"):
                # 提取 TITLE- 后的标题类型（如"规则"、"奖励"）
                title_type = v.split("TITLE-")[-1].strip()
                # 读取同行右侧（即下一列）的文本作为实际标题
                title_text = row_vals[i + 1] if i + 1 < len(row_vals) else ""
                titles.append((title_type, title_text, r))
                break
    return titles


def parse_section_block(ws, c1, c2, r_start, r_end, merge_idx):
    sections = []
    r = r_start
    block_free_paragraphs = []  # 改为段落数组

    def collect_rewards(r0, section_title):
        items = []
        rr = r0
        while rr <= r_end:
            rv = row_region_values(ws, rr, c1, c2, merge_idx)
            first = rv[0] if rv else ""
            # stop if next block begins (but allow first line)
            if first.startswith("TITLE-") or first.startswith("RINK-") or first.startswith("RANK-") or first.startswith("RULES-") or first.startswith("TABLE-"):
                if rr != r0:
                    break
            if is_row_blank(rv):
                if rr != r0:
                    break
            name = rv[1] if len(rv) > 1 else ""
            img_text = rv[2] if len(rv) > 2 else ""
            desc = rv[3] if len(rv) > 3 else ""
            if not name:
                for vv in rv[1:]:
                    if vv:
                        if not name:
                            name = vv
                        elif not desc:
                            desc = vv
            if name or desc or img_text:
                items.append({
                    "name": name,
                    "image": img_text,  # 原始单元格文本（期望文件名）
                    "desc": desc,
                    "_row": rr,
                    "_img_col": c1 + 2,
                    "_expected": img_text or name,
                })
            rr += 1
        sections.append({"title": section_title, "content": "", "rewards": items})
        return rr

    def collect_table(r0):
        """
        收集表格数据
        TABLE- 标记后的所有行都是表格内容（没有表头概念）
        直到下一个 TABLE- 或其他标记为止
        支持文字和图片、居中、加粗等格式
        
        重要：使用 TABLE- 标记所在行的合并单元格范围作为表格列范围
        """
        table_data = {
            "type": "table",
            "rows": []
        }
        
        # 查找 TABLE- 标记的实际列范围（使用合并单元格范围）
        table_c1 = c1
        table_c2 = c2
        
        # 检查 TABLE- 标记所在的单元格是否有合并
        rv_table_marker = row_region_values(ws, r0, c1, c2, merge_idx)
        for col_idx, val in enumerate(rv_table_marker):
            if val.startswith("TABLE-"):
                actual_col = c1 + col_idx
                # 检查该单元格是否在合并区域中
                if (r0, actual_col) in merge_idx:
                    min_r, min_c, max_r, max_c = merge_idx[(r0, actual_col)]
                    # 使用合并区域的列范围（跳过标记列本身，从下一列开始）
                    table_c1 = min_c + 1 if min_c == c1 else min_c
                    table_c2 = max_c
                else:
                    # 如果 TABLE- 在第一列，表格从第二列开始到 region 结束
                    table_c1 = c1 + 1
                    table_c2 = c2
                break
        
        rr = r0 + 1  # 跳过 TABLE- 标记行，从下一行开始是表格内容
        
        # 所有行都是数据行（没有表头）
        while rr <= r_end:
            # 检查标记列（c1）是否有新的块级标记
            marker_col_val = clean_text(get_value(ws, rr, c1, merge_idx))
            if marker_col_val and (marker_col_val.startswith("TITLE-") or marker_col_val.startswith("RINK-") or marker_col_val.startswith("RANK-")):
                break
            
            # 使用表格的实际列范围获取数据
            rv = row_region_values(ws, rr, table_c1, table_c2, merge_idx)
            
            # 检查内容列是否有 TABLE- 结束标记
            has_table_in_content = any(val.startswith("TABLE-") for val in rv if val)
            if has_table_in_content:
                break
            
            # 检查是否是空行（但不要停止，要保留空行）
            # 只有当连续遇到多个空行或到达表格末尾标记时才停止
            is_empty_row = is_row_blank(rv)
            
            # 如果是空行，检查下一行是否也是空行或有标记
            # 连续两个空行才停止（避免表格中间的单个空行被跳过）
            if is_empty_row and rr + 1 <= r_end:
                next_marker = clean_text(get_value(ws, rr + 1, c1, merge_idx))
                next_rv = row_region_values(ws, rr + 1, table_c1, table_c2, merge_idx)
                # 如果下一行是空行或者标记列有标记，则停止
                if is_row_blank(next_rv) or (next_marker and (next_marker.startswith("TITLE-") or next_marker.startswith("RINK-") or next_marker.startswith("RANK-"))):
                    break
            
            # 收集这一行的数据（表格范围内的所有列）
            row_data = []
            col_output_idx = 0  # 输出列索引
            
            for col_idx, val in enumerate(rv):
                # 检查当前单元格是否在合并区域内
                actual_row = rr
                actual_col = table_c1 + col_idx  # 使用表格的起始列
                rowspan = 1
                colspan = 1
                is_merged_child = False
                
                if (actual_row, actual_col) in merge_idx:
                    min_r, min_c, max_r, max_c = merge_idx[(actual_row, actual_col)]
                    rowspan = max_r - min_r + 1
                    colspan = max_c - min_c + 1
                    # 如果不是合并区域的左上角单元格，标记为合并子单元格（不输出）
                    if actual_row != min_r or actual_col != min_c:
                        is_merged_child = True
                
                # 合并区域的子单元格跳过（前端会自动处理）
                if is_merged_child:
                    continue
                
                # 获取单元格格式信息（加粗、居中）
                cell_info = get_cell_info(ws, actual_row, actual_col, merge_idx)
                is_bold = cell_info['bold']
                is_center = cell_info['center']
                
                # 检测是否是图片标记（包含文件扩展名或 image: 前缀）
                is_image = False
                image_id = None
                
                if val:
                    val_lower = val.lower()
                    # 检测图片格式
                    if any(ext in val_lower for ext in ['.png', '.jpg', '.jpeg', '.gif', '.webp']) or val_lower.startswith('image:'):
                        is_image = True
                        # 记录图片位置信息用于后续图片匹配
                        image_id = val if not val_lower.startswith('image:') else val[6:].strip()
                
                cell_data = {
                    "value": val if val else "",
                    "is_image": is_image,
                    "rowspan": rowspan,
                    "colspan": colspan,
                    "bold": is_bold,
                    "center": is_center,
                    "_row": actual_row,
                    "_col": actual_col,
                    "_expected": image_id if is_image else None
                }
                row_data.append(cell_data)
                col_output_idx += 1
            
            # 始终添加行数据，即使为空（空行或被合并单元格占用的行）
            # 这样可以保持行号的连续性，前端能正确处理 rowspan 和显示空行
            table_data["rows"].append(row_data)
            
            rr += 1
        
        # 将表格作为特殊的 section 添加
        sections.append({
            "title": "",
            "content": "",
            "rewards": [],
            "table": table_data
        })
        
        # 如果当前行的内容列包含 TABLE- 结束标记，跳过它
        if rr <= r_end:
            rv_check = row_region_values(ws, rr, table_c1, table_c2, merge_idx)
            has_table_end = any(val.startswith("TABLE-") for val in rv_check if val)
            if has_table_end:
                rr += 1
        
        return rr

    def collect_rules_content(r0, section_title):
        """
        收集规则内容，返回结构化的段落数组
        
        每个 Excel 行 → 一个 Paragraph
        每个单元格（或合并单元格） → 一个 TextRun
        
        只保留：文本、对齐（左中右）、加粗、斜体、颜色
        """
        paragraphs = []
        seen_lines = set()  # 用于跨行去重
        rr = r0

        while rr <= r_end:
            # 读取当前行的所有单元格（带格式）
            cells_info = []
            for c in range(c1, c2 + 1):
                cell_info = get_cell_info(ws, rr, c, merge_idx)
                cells_info.append(cell_info)
            
            # 提取纯文本用于检查
            rv = [clean_text(cell['value']) for cell in cells_info]

            # 如果不是起始行，检查是否遇到标记或空行
            if rr > r0:
                first = rv[0] if rv else ""
                if first and (first.startswith("TITLE-") or first.startswith("RINK-") or first.startswith("RANK-")):
                    break
                
                has_table_in_content = any(val.startswith("TABLE-") for val in rv[1:] if val)
                if has_table_in_content:
                    break
                
                # 如果是空行，检查下一行是否也是空行
                # 连续两个空行才停止（允许单个空行）
                if is_row_blank(rv):
                    if rr + 1 <= r_end:
                        next_rv = row_region_values(ws, rr + 1, c1, c2, merge_idx)
                        next_first = next_rv[0] if next_rv else ""
                        # 如果下一行也是空行或者是标记行，则停止
                        if is_row_blank(next_rv) or next_first.startswith("TITLE-") or next_first.startswith("RINK-") or next_first.startswith("RANK-"):
                            break
                    else:
                        # 已经是最后一行了，停止
                        break

            # 构建段落
            paragraph_align = "left"  # 默认左对齐
            runs = []
            
            # 列去重并构建 runs（跳过第一列标记列）
            seen_row = set()
            for cell_info in cells_info[1:]:  # 跳过第一列
                val = clean_text(cell_info['value'])
                if not val or val in seen_row:
                    continue
                
                seen_row.add(val)
                
                # 构建 TextRun（只包含需要的字段）
                run = {"text": val}
                
                # 添加样式（只在非默认值时添加）
                if cell_info.get('bold'):
                    run['bold'] = True
                if cell_info.get('italic'):
                    run['italic'] = True
                if cell_info.get('color'):
                    run['color'] = cell_info['color']
                
                # 判断段落对齐方式（从第一个有效单元格读取）
                alignment = cell_info.get('alignment')
                if alignment == 'center':
                    paragraph_align = 'center'
                elif alignment == 'right':
                    paragraph_align = 'right'
                elif alignment == 'left':
                    paragraph_align = 'left'
                
                runs.append(run)
            
            # 行去重：如果这一行的内容之前已经出现过，跳过
            # 但空行总是保留（不参与去重）
            if runs:
                # 用所有 runs 的文本组合作为去重 key
                line_key = '|'.join(run['text'] for run in runs)
                if line_key not in seen_lines:
                    paragraphs.append({
                        "align": paragraph_align,
                        "runs": runs
                    })
                    seen_lines.add(line_key)
            else:
                # 空行：添加一个空 paragraph（runs 为空数组，用特殊文本表示空行）
                paragraphs.append({
                    "align": "left",
                    "runs": [{"text": ""}]  # 空文本表示空行
                })
            
            rr += 1

        # 返回段落数组
        sections.append({
            "title": section_title,
            "paragraphs": paragraphs,
            "rewards": []
        })
        return rr

    while r <= r_end:
        rv = row_region_values(ws, r, c1, c2, merge_idx)
        if is_row_blank(rv):
            r += 1
            continue
        
        # 检查第一列（标记列）的标记类型
        first = rv[0] if rv else ""
        
        # 检查内容列（rv[1:]）是否有 TABLE- 标记
        has_table_in_content = any(val.startswith("TABLE-") for val in rv[1:] if val)
        if has_table_in_content:
            r = collect_table(r)
            continue
        
        if first.startswith("RULES-"):
            title = first.split("RULES-")[-1].strip()
            # 找到具有相同标题的所有连续行的结束位置
            rules_end = r
            while rules_end + 1 <= r_end:
                next_rv = row_region_values(ws, rules_end + 1, c1, c2, merge_idx)
                next_first = next_rv[0] if next_rv else ""

                # 如果遇到空行，检查是否是内容末尾
                if is_row_blank(next_rv):
                    break

                # 如果下一行是 TITLE 或 RINK/RANK 或 TABLE 标记，停止
                if next_first.startswith("TITLE-") or next_first.startswith("RINK-") or next_first.startswith("RANK-") or next_first.startswith("TABLE-"):
                    break

                # 如果下一行也是 RULES 标记
                if next_first.startswith("RULES-"):
                    next_title = next_first.split("RULES-")[-1].strip()
                    if next_title == title:
                        # 相同标题，继续合并（这是合并单元格导致的重复）
                        rules_end += 1
                    else:
                        # 不同标题，停止
                        break
                else:
                    # 非标记行，属于当前 RULES 的内容
                    rules_end += 1

            # 使用修改后的范围解析内容
            temp_r_end = r_end
            r_end = rules_end
            r = collect_rules_content(r, title)
            r_end = temp_r_end
            continue
        # 支持 RANK- 和 RINK-（兼容旧标记）
        if first.startswith("RANK-") or first.startswith("RINK-"):
            if first.startswith("RANK-"):
                sub = first.split("RANK-")[-1].strip()
            else:
                sub = first.split("RINK-")[-1].strip()
            r = collect_rewards(r, sub)
            continue
        # fallback 分支：构建结构化段落
        # 读取当前行的所有单元格（带格式）
        cells_info = []
        for c in range(c1, c2 + 1):
            cell_info = get_cell_info(ws, r, c, merge_idx)
            cells_info.append(cell_info)
        
        # 构建段落
        paragraph_align = "left"
        runs = []
        seen_row = set()
        
        for cell_info in cells_info[1:]:  # 跳过第一列（标记列）
            val = clean_text(cell_info['value'])
            if not val or val in seen_row:
                continue
            
            seen_row.add(val)
            
            # 构建 TextRun
            run = {"text": val}
            
            # 添加样式
            if cell_info.get('bold'):
                run['bold'] = True
            if cell_info.get('italic'):
                run['italic'] = True
            if cell_info.get('color'):
                run['color'] = cell_info['color']
            
            # 判断段落对齐方式
            alignment = cell_info.get('alignment')
            if alignment == 'center':
                paragraph_align = 'center'
            elif alignment == 'right':
                paragraph_align = 'right'
            
            runs.append(run)
        
        # 添加段落（包括空行）
        if runs:
            block_free_paragraphs.append({
                "align": paragraph_align,
                "runs": runs
            })
        else:
            # 空行：添加一个空 paragraph
            block_free_paragraphs.append({
                "align": "left",
                "runs": [{"text": ""}]  # 空文本表示空行
            })
        
        r += 1

    # 返回段落数组
    return sections, block_free_paragraphs


def merge_reward_sections(sections):
    merged = []
    title_to_index = {}
    for s in sections:
        if isinstance(s, dict) and s.get("rewards"):
            title = s.get("title", "")
            if title in title_to_index:
                idx = title_to_index[title]
                merged[idx]["rewards"].extend(s.get("rewards") or [])
            else:
                title_to_index[title] = len(merged)
                merged.append({
                    "title": title,
                    "paragraphs": [],  # 使用空段落数组而不是空字符串
                    "rewards": list(s.get("rewards") or [])
                })
        else:
            merged.append(s)
    return merged


def parse_sheet(ws) -> dict:
    """
    解析工作表，按 TITLE 分组生成 blocks 结构。

    返回格式:
    {
      "pages": [
        {
          "region": "TW",
          "blocks": [
            {
              "block_title": "活动规则说明",     # TITLE- 右侧的文本
              "block_type": "rules",             # 根据 TITLE- 判断：规则/奖励
              "sections": [...]                  # 该分组下的所有内容
            },
            {
              "block_title": "真爱树榜单奖励",
              "block_type": "rewards",
              "sections": [...]
            }
          ]
        }
      ]
    }
    """
    merge_idx = build_merge_index(ws)
    regions = find_regions(ws, merge_idx)
    pages = []
    for region in regions:
        c1, c2 = region["col_start"], region["col_end"]
        titles = scan_titles(ws, region, merge_idx)
        blocks = []

        for i, (title_type, title_text, trow) in enumerate(titles):
            r_start = trow + 1
            r_end = titles[i + 1][2] - 1 if i + 1 < len(titles) else ws.max_row
            secs, fallback = parse_section_block(ws, c1, c2, r_start, r_end, merge_idx)

            if secs:
                merged_secs = merge_reward_sections(secs)
            else:
                # fallback 现在返回段落数组
                merged_secs = [{"title": title_text or title_type, "paragraphs": fallback, "rewards": []}]

            # 智能判断 block 类型（基于内容和标题关键词）
            # 1. 检查是否包含奖励数据
            total_rewards = sum(len(sec.get("rewards", [])) for sec in merged_secs)
            
            # 2. 检查标题关键词
            title_lower = title_type.lower()
            reward_keywords = ["奖励", "reward", "榜", "排行", "排名", "leaderboard", "prize", "奖品"]
            has_reward_keyword = any(kw in title_lower for kw in reward_keywords)
            
            # 3. 综合判断：有奖励数据 或 标题包含关键词
            block_type = "rewards" if (total_rewards > 0 or has_reward_keyword) else "rules"

            # 确定最终的标题文本，并去掉可能残留的 TITLE- 前缀
            final_title = title_text or title_type
            # 安全检查：去掉任何残留的 TITLE- 前缀（以防格式不标准）
            if final_title.startswith("TITLE-"):
                final_title = final_title.split("TITLE-", 1)[-1].strip()
            
            blocks.append({
                "block_title": final_title,
                "block_type": block_type,
                "sections": merged_secs
            })

        # 根据地区代码判断文本方向
        direction = "rtl" if is_rtl_region(region["code"]) else "ltr"
        
        pages.append({
            "region": region["code"],
            "direction": direction,
            "blocks": blocks
        })
    return {"pages": pages}


def has_region_marker(ws) -> bool:
    """检查工作表第一行是否包含 REGION- 标记"""
    if ws.max_row < 1:
        return False
    
    for c in range(1, ws.max_column + 1):
        cell_value = ws.cell(1, c).value
        if cell_value and str(cell_value).strip().startswith("REGION-"):
            return True
    return False


def parse_file(xlsx_path: str, sheet: Optional[str] = None) -> dict:
    """
    统一的多 sheet 解析函数
    
    - 如果指定 sheet 参数，只解析该 sheet（如果有 REGION-）
    - 如果不指定，解析所有包含 REGION- 的 sheet
    
    返回格式：
    {
        "sheets": {
            "Sheet1": { "pages": [...] },
            "Sheet2": { "pages": [...] }
        },
        "skipped_sheets": ["Sheet3"]
    }
    """
    wb = load_workbook(xlsx_path, data_only=True)
    sheets_result = {}
    skipped = []
    
    if sheet:
        # 指定了 sheet，只解析这一个
        if sheet in wb.sheetnames:
            ws = wb[sheet]
            if has_region_marker(ws):
                sheets_result[sheet] = parse_sheet(ws)
            else:
                skipped.append(sheet)
        else:
            raise ValueError(f"Sheet '{sheet}' not found")
    else:
        # 未指定，解析所有有 REGION- 的 sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if has_region_marker(ws):
                sheets_result[sheet_name] = parse_sheet(ws)
            else:
                skipped.append(sheet_name)
    
    return {
        "sheets": sheets_result,
        "skipped_sheets": skipped
    }

